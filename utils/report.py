import os
import shutil
from datetime import datetime

import openpyxl
from babel.dates import format_date
from num2words import num2words
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.cell.cell import MergedCell

from db.models.models import Check, User


def set_cell_value(sheet, row, col, value):
    """
    Безопасная запись в ячейку Excel, учитывая объединенные ячейки.
    Если ячейка объединена, записывает значение в верхнюю левую ячейку диапазона.
    """
    cell = sheet.cell(row=row, column=col)
    
    if isinstance(cell, MergedCell):
        for merged_range in sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                top_left_cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                top_left_cell.value = value
                return
    else:
        cell.value = value


async def create_report(user_tg_id: int) -> str:
    user = await User.get(tg_id=user_tg_id)
    checks = await Check.filter(user_id=user.id, used=False)

    file_path = f'data/{user.id}.xlsx'
    shutil.copy2('data/report.xlsx', file_path)

    workbook = openpyxl.load_workbook(filename=file_path)
    sheet = workbook.active

    now = datetime.now()
    amount = sum([check.sum for check in checks]) / 100
    today = datetime.now().strftime('%d.%m.%Y')
    name = short_name(user.full_name)
    boss_name = short_name(user.boss_name)
    rub_amount = format_amount(amount)
    cop_amount = str(amount).split('.')[1]
    text_amount = f"{num2words(int(amount), lang='ru')} рублей {cop_amount} копеек ({rub_amount} руб. {cop_amount} коп.)"

    # Вставка данных
    set_cell_value(sheet, 9, 18, rub_amount)
    set_cell_value(sheet, 9, 24, cop_amount)
    set_cell_value(sheet, 13, 10, today)
    set_cell_value(sheet, 15, 15, today)
    set_cell_value(sheet, 19, 8, user.subdivision)
    set_cell_value(sheet, 21, 6, name)
    set_cell_value(sheet, 23, 17, 'Расходы ' + format_date(now, format='MMMM yyyy', locale='ru'))
    set_cell_value(sheet, 33, 10, f'{rub_amount},{cop_amount}')
    set_cell_value(sheet, 39, 10, text_amount)
    set_cell_value(sheet, 55, 9, name)
    set_cell_value(sheet, 56, 4, format_date(now, format='d MMMM yyyy г.', locale='ru'))
    set_cell_value(sheet, 56, 11, text_amount)

    # Создание таблицы чеков
    start_row, start_column, end_row, end_column = 66, 2, 66, 24
    source_row = [cell.value for cell in sheet[66]]
    sheet.delete_rows(66)

    for i, check in enumerate(checks):
        end_row = 66 + i
        sheet.insert_rows(end_row)

        for col_num, value in enumerate(source_row, start=1):
            sheet.cell(row=end_row, column=col_num, value=value)

        float_sum = check.sum / 100
        check_rub_amount = format_amount(float_sum)
        check_cop_amount = str(float_sum).split('.')[1]

        set_cell_value(sheet, end_row, 2, i + 1)
        set_cell_value(sheet, end_row, 4, check.date.strftime('%d.%m.%Y'))
        set_cell_value(sheet, end_row, 6, check.id)
        set_cell_value(sheet, end_row, 8, check.type)
        set_cell_value(sheet, end_row, 12, f'{check_rub_amount},{check_cop_amount}')
        set_cell_value(sheet, end_row, 18, f'{check_rub_amount},{check_cop_amount}')

    table_summ_row = end_row + 1
    set_cell_value(sheet, table_summ_row, 12, f'{rub_amount},{cop_amount}')
    set_cell_value(sheet, table_summ_row, 18, f'{rub_amount},{cop_amount}')

    user_signature_row = end_row + 3
    set_cell_value(sheet, user_signature_row, 2, 'Подотчетное лицо')
    set_cell_value(sheet, user_signature_row, 14, name)

    boss_signature_row = end_row + 7
    set_cell_value(sheet, boss_signature_row, 2, 'Руководитель')
    set_cell_value(sheet, boss_signature_row, 14, boss_name)

    workbook.save(file_path)

    for check in checks:
        await check.update(used=True)

    return file_path


def format_amount(amount: float):
    return f"{int(str(amount).split('.')[0]):,}".replace(",", " ")


def short_name(full_name: str) -> str:
    name_parts = full_name.split()
    return f"{name_parts[0]} {name_parts[1][0]}. {name_parts[2][0]}."
